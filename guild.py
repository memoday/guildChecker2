import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import os, sys
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
from PyQt5 import uic
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import checkInfo as ci
import tracker as track
import webbrowser

__version__ = "DEV"

header = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Whale/3.18.154.13 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
}

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

icon = resource_path('assets/memo.ico')
form = resource_path('ui/main.ui')

form_class = uic.loadUiType(form)[0]

wb = openpyxl.Workbook()
sheet = wb.active
membersList = []

def selectedWorld(worldName):
    world = [None,'리부트','리부트2','오로라','레드','이노시스','유니온','스카니아','루나','제니스','크로아','베라','엘리시움','아케인','노바']
    worldIndex = world.index(worldName)
    return worldIndex

def fileCreate(serverName, guildName):
    global ws1,wb,fileName
    now = datetime.datetime.now()
    
    now = now.strftime('%Y-%m-%d')

    fileName = serverName+'_'+guildName+'_'+now+'.xlsx'
    try:
        wb = openpyxl.load_workbook(fileName)
        print('file exists')
        return False
    except FileNotFoundError:
        print('File Not Found:')
        wb = openpyxl.Workbook()
    ws1 = wb.active

def tempFileCreate():

    global tempx, tempwb

    tempwb = openpyxl.Workbook()
    tempx = tempwb.active

def tempGuildCrawl(driver):

    page = 1
    membersCount = 0

    nowURL = driver.current_url

    global guildlist
    guildlist = []


    while True:

        newURL = nowURL+'&orderby=0&page='+str(page)+''
        raw = requests.get(newURL,headers=header)
        html = BeautifulSoup(raw.text,"html.parser")   
        members = html.select('#container > div > div > table > tbody > tr')
        
        for member in range(20):
            membersCount += 1
            nickName = members[member].select_one('td.left > dl > dt > a').text
            print(nickName)
            guildlist.append(nickName)
        page += 1
        time.sleep(1)

def guildCrawl(driver,self):

    page = 1
    membersCount = 0

    nowURL = driver.current_url

    while True:

        newURL = nowURL+'&orderby=0&page='+str(page)+''
        raw = requests.get(newURL,headers=header)
        html = BeautifulSoup(raw.text,"html.parser")   
        members = html.select('#container > div > div > table > tbody > tr')
        
        for member in range(20):
            membersCount += 1
            nickName = members[member].select_one('td.left > dl > dt > a').text
            print(nickName)
            ws1.append([nickName])
            self.statusBar().showMessage(nickName)
        page += 1
        time.sleep(1)

def finalCheck(self, guildName):
    set1 = set(guildlist)
    set2 = set(oldGuildList)
    changeCount = 0
    
    guildIn = list(set1 - set2)
    guildOut = list(set2 - set1)

    newList = []
    trackList = [] #닉변 추적 기능 임시 비활성화, 기존 길드 = 변경 길드 일 경우 찾을 수 없음으로 표기
    errorList = [] #기존 trackList를 errorList로 옮김
    leaveList = []
    transferList = []


    for i in range(len(guildIn)):
        print('[신규]',guildIn[i])
        changed = ('[신규] '+guildIn[i])
        self.guildMembers_changed.append(changed)
        newList.append(guildIn[i])
        changeCount += 1


    for i in range(len(guildOut)):
        time.sleep(0.3)
        check, newGuild = ci.checkGuild(guildOut[i], guildName)
        if newGuild == '':
            print('[탈퇴]', guildOut[i])
            changed = ('[탈퇴] '+guildOut[i])
            leaveList.append(changed)
            changeCount += 1
        elif newGuild == guildName:
            # print('[닉변/캐삭]',guildOut[i])
            # changed = ('[닉변/캐삭] '+guildOut[i])
            # self.guildMembers_changed.append(changed)
            notFound = ('[확인불가] '+guildOut[i])
            errorList.append(notFound)
            changeCount += 1
        else:
            print('[이전]', guildOut[i],'-> ',newGuild)
            changed = ('[이전] '+guildOut[i]+' -> '+newGuild)
            transferList.append(changed)
            changeCount += 1

    for i in range(len(leaveList)):
        self.guildMembers_changed.append(leaveList[i])
    
    for i in range(len(transferList)):
        self.guildMembers_changed.append(transferList[i])

    nickChangeResult = []
    unverifiedResult = []

    for i in range(len(trackList)): #닉변 확인

        changed_to = track.tracker(trackList[i],newList)

        if len(changed_to) > 0:
            result = '/'.join(changed_to)
            changed = '[닉변]'+trackList[i]+' -> '+result
            nickChangeResult.append(changed)

        else:
            changed = ('[확인불가]'+trackList[i])
            unverifiedResult.append(changed)

    for i in range(len(errorList)):
        self.guildMembers_changed.append(errorList[i])
    
    for i in range(len(nickChangeResult)):
        self.guildMembers_changed.append(nickChangeResult[i])
    
    for i in range(len(unverifiedResult)):
        self.guildMembers_changed.append(unverifiedResult[i])
        
    self.changeCount.setText(str(changeCount)+' 명')

class execute(QThread):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        
    def run(self):

        self.parent.btn_start.setDisabled(True)
        self.parent.statusBar().showMessage('길드원 추출 준비 중..')

        worldName = str(self.parent.combo_serverName.currentText())
        worldNumber = selectedWorld(worldName)

        guildName = self.parent.input_guildName.text()
        if guildName == "":
            self.parent.statusBar().showMessage('추출하기: 길드 이름을 입력해주세요')
            return

        url = f"https://maplestory.nexon.com/Ranking/World/Guild?w={worldNumber}&t=1&n={guildName}"

        try:
            raw = requests.get(url,headers=header)
            html = BeautifulSoup(raw.text,"html.parser")
            href = html.select_one('#container > div > div > div:nth-child(4) > div.rank_table_wrap > table > tbody > tr > td:nth-child(2) > span > a')['href']

            guildUrl = 'https://maplestory.nexon.com'+href
            pageNumber = 1

            while True:
                self.parent.statusBar().showMessage(f'{guildName} 길드원 추출 중: {pageNumber} /10')
                guildUrlPage = guildUrl+f'&page={pageNumber}'
                try:
                    self.crawlMembers(guildUrlPage)
                    print('크롤링한 페이지: ',pageNumber)
                except Exception as e:
                    print(e)
                    break

                if pageNumber == 10:

                    print('10번째 페이지로 크롤링을 종료합니다\n')
                    print(membersList)
                    print(len(membersList))

                    for i in range(len(membersList)):
                        sheet.append(membersList[i])
                    self.parent.guildMembers.append(membersList)
                    now = datetime.datetime.now()
                    now = now.strftime('%Y-%m-%d')
                    wb.save(f"{worldName}_{guildName}_{now}.csv")
                    self.parent.count.setText(str(len(membersList))+' 명')
                    self.parent.statusBar().showMessage('추출하기 완료. '+guildName)
                    self.parent.btn_start.setEnabled(True)

                    break
                else:
                    pageNumber += 1

        except Exception as e:
            self.parent.statusBar().showMessage(f'[ERROR] {e}')
        
    def crawlMembers(self,guildUrl):
        r = requests.get(guildUrl,headers=header)
        html = BeautifulSoup(r.text,"html.parser") 
        members = html.select('#container > div > div > table > tbody > tr')
        for member in members:
            nick = member.select_one('td.left > span > img')['alt']
            # job = member.select_one('td.left > dl > dd').text #일부 직업들은 기사단/마법사/전사/해적 등으로 표시되어있음
            membersList.append([nick])

       
class WindowClass(QMainWindow, form_class):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        #프로그램 기본설정
        self.setWindowIcon(QIcon(icon))
        self.setWindowTitle('Guild Checker '+__version__)
        self.statusBar().showMessage('프로그램 정상 구동 중')

        #실행 후 기본값 설정

        #버튼 기능
        self.btn_start.clicked.connect(self.main)
        self.btn_exit.clicked.connect(self.exit)
        self.btn_load.clicked.connect(self.fileLoad)
        self.btn_check.clicked.connect(self.checkInfo)
        self.btn_github.clicked.connect(self.github)
    
    def main(self):
        self.guildMembers_changed.setText('')
        x = execute(self)
        x.start()

    def fileLoad(self): #파일 불러오기
        global sheet, oldGuildList
        fname = QFileDialog.getOpenFileName(self,'','','Excel(*.xlsx) ;;All File(*)')
        
        self.guildMembers_changed.setText('')
        self.changeCount.setText('- 명')

        loadedFile = QFileInfo(fname[0]).fileName()
        if loadedFile != "":
            self.statusBar().showMessage('파일을 불러왔습니다. '+loadedFile)

        try:
            loadedFileServer, loadedFileGuild, loadedFileDate = loadedFile.split('_')
            self.input_guildName.setText(loadedFileGuild)
            self.combo_serverName.setCurrentText(loadedFileServer)
        except ValueError:
            pass

        count = 0
        oldGuildList = []
        if fname[0]:
            f = open(fname[0],'r')
            data = openpyxl.load_workbook(filename= fname[0],data_only=True)
            sheet = data['Sheet']
            
            try:
                for i in list(sheet.columns)[0]:
                    count += 1
                    self.guildMembers.append(i.value)
                    oldGuildList.append(i.value)
                self.count.setText(str(count)+' 명')
            except IndexError:
                self.statusBar().showMessage('불러올 길드원이 없습니다. '+loadedFile)

    def checkInfo(self): #변동사항 확인
        print('')

    def github(self):
        webbrowser.open_new_tab('https://github.com/memoday/guildMemberChecker/releases')

    def exit(self):
        os.system("taskkill /f /im chromedriver.exe") #chomrdriver.exe 강제종료
        sys.exit(0)

    def closeEvent(self, event):
        os.system("taskkill /f /im chromedriver.exe") #chomrdriver.exe 강제종료
        sys.exit(0)

if __name__ == "__main__":
    app = QApplication(sys.argv) 
    myWindow = WindowClass() 
    myWindow.show()
    app.exec_()
